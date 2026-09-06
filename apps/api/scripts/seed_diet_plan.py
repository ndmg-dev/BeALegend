"""Semeia o plano alimentar a partir da planilha de dieta, para um usuário existente.

    python scripts/seed_diet_plan.py --email voce@exemplo.com

Roda como a role OWNER (bypassa RLS), porque cria o catálogo global de
alimentos e de suplementos (``is_global=true``, ``user_id NULL``), o que a
role de runtime não pode fazer.

Um plano de cada vez fica ativo: ao criar o novo, os planos alimentares
ativos anteriores passam a ``ativo=false`` (continuam no histórico — os
registros de refeição apontam para os slots antigos e não podem sumir).

Idempotente: rodar de novo sem ``--force`` atualiza o catálogo global e pula
a criação do plano se já existir um com o mesmo nome. Com ``--force``, o
plano antigo e seus slots são apagados e recriados — note que isso zera o
``slot_id`` dos ``meal_log`` que apontavam para eles (ON DELETE SET NULL): o
registro do que foi comido sobrevive, o vínculo com a refeição planejada não.

Parser preso ao formato desta planilha (abas Painel, Base alimentos, Plano
diário, Suplementação). Ver ``app/seed/diet_parsing.py`` para o que fica de
fora e por quê.
"""

from __future__ import annotations

import argparse
import asyncio
import sys
from pathlib import Path
from uuid import UUID

import openpyxl
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.ids import uuid7  # noqa: E402
from app.models import (  # noqa: E402
    FoodItem,
    MealPlan,
    MealSlot,
    MealSlotItem,
    NutritionTarget,
    Supplement,
    User,
)
from app.seed.diet_parsing import (  # noqa: E402
    AlimentoPlanilha,
    MetaPlanilha,
    RefeicaoPlanilha,
    SuplementoPlanilha,
    parse_alimentos,
    parse_meta,
    parse_refeicoes,
    parse_suplementos,
)

DEFAULT_XLSX = (
    Path(__file__).resolve().parent.parent / "data" / ("planilha_dieta_massa_definicao.xlsx")
)
DEFAULT_PLAN_NAME = "Ganho de massa com controle de gordura"

#: Horário sugerido por refeição. A planilha não traz horário — o app usa o
#: campo para ordenar e lembrar, então um palpite explícito aqui é melhor do
#: que deixar a coluna vazia. Casa por prefixo do nome da refeição.
_HORARIOS = {
    "café": "07:30",
    "almoço": "12:30",
    "lanche": "16:00",
    "jantar": "20:00",
    "ceia": "22:00",
}


def _ler_linhas(ws, colunas_esperadas: list[str]) -> list[dict]:
    """Lê uma aba em lista de dicts, usando a linha de cabeçalho real (a
    primeira que contém todas as colunas esperadas)."""
    linhas = list(ws.iter_rows(values_only=True))
    cabecalho_idx = next(
        i
        for i, linha in enumerate(linhas)
        if linha and all(col in (linha or ()) for col in colunas_esperadas)
    )
    cabecalho = linhas[cabecalho_idx]
    dados = []
    for linha in linhas[cabecalho_idx + 1 :]:
        if not linha or all(celula is None for celula in linha):
            continue
        dados.append(dict(zip(cabecalho, linha, strict=False)))
    return dados


def carregar_planilha(caminho: Path) -> dict:
    wb = openpyxl.load_workbook(caminho, data_only=True)
    return {
        "alimentos": parse_alimentos(_ler_linhas(wb["Base alimentos"], ["Alimento", "kcal"])),
        "refeicoes": parse_refeicoes(_ler_linhas(wb["Plano diário"], ["Refeição", "Alimento"])),
        "suplementos": parse_suplementos(
            _ler_linhas(wb["Suplementação"], ["Item", "Como usar na planilha"])
        ),
        "meta": parse_meta(list(wb["Painel"].iter_rows(values_only=True))),
    }


async def _obter_usuario(session: AsyncSession, email: str) -> User:
    usuario = await session.scalar(select(User).where(User.email == email.lower()))
    if usuario is None:
        raise SystemExit(
            f"Nenhum usuário com o e-mail {email!r}. Cadastre-se no app antes de rodar o seed."
        )
    return usuario


async def _upsert_alimentos_globais(
    session: AsyncSession, alimentos: list[AlimentoPlanilha]
) -> dict[str, UUID]:
    """Catálogo global: is_global=true, user_id NULL. Visível a todos."""
    ids: dict[str, UUID] = {}
    for alimento in alimentos:
        existente = await session.scalar(
            select(FoodItem).where(FoodItem.is_global.is_(True), FoodItem.nome == alimento.nome)
        )
        alvo = existente or FoodItem(id=uuid7(), user_id=None, is_global=True, nome=alimento.nome)
        alvo.kcal = alimento.kcal
        alvo.proteina_g = alimento.proteina_g
        alvo.carboidrato_g = alimento.carboidrato_g
        alvo.gordura_g = alimento.gordura_g
        alvo.fibra_g = alimento.fibra_g
        alvo.referencia_pratica = alimento.referencia_pratica
        alvo.fonte = alimento.fonte
        alvo.conferir_rotulo = alimento.conferir_rotulo
        if existente is None:
            session.add(alvo)
            await session.flush()
        ids[alimento.nome] = alvo.id
    return ids


async def _upsert_suplementos_globais(
    session: AsyncSession, suplementos: list[SuplementoPlanilha]
) -> None:
    for suplemento in suplementos:
        existente = await session.scalar(
            select(Supplement).where(
                Supplement.is_global.is_(True), Supplement.nome == suplemento.nome
            )
        )
        alvo = existente or Supplement(
            id=uuid7(), user_id=None, is_global=True, nome=suplemento.nome
        )
        alvo.como_usar = suplemento.como_usar
        alvo.faixa = suplemento.faixa
        alvo.horario = suplemento.horario
        alvo.observar = suplemento.observar
        alvo.fonte = suplemento.fonte
        alvo.status = suplemento.status
        alvo.ordem = suplemento.ordem
        if existente is None:
            session.add(alvo)
            await session.flush()


def _horario_sugerido(nome_refeicao: str) -> str | None:
    chave = nome_refeicao.strip().lower()
    for prefixo, horario in _HORARIOS.items():
        if chave.startswith(prefixo):
            return horario
    return None


async def _criar_plano(
    session: AsyncSession,
    usuario: User,
    nome_plano: str,
    refeicoes: list[RefeicaoPlanilha],
    alimentos_por_nome: dict[str, UUID],
    meta: MetaPlanilha,
    force: bool,
) -> None:
    existente = await session.scalar(
        select(MealPlan).where(
            MealPlan.user_id == usuario.id,
            MealPlan.nome == nome_plano,
            MealPlan.deleted_at.is_(None),
        )
    )
    if existente is not None:
        if not force:
            print(f"Plano {nome_plano!r} já existe para {usuario.email}. Use --force para recriar.")
            return
        await session.execute(delete(MealPlan).where(MealPlan.id == existente.id))
        await session.flush()

    anteriores = (
        await session.scalars(
            select(MealPlan).where(
                MealPlan.user_id == usuario.id,
                MealPlan.ativo.is_(True),
                MealPlan.deleted_at.is_(None),
            )
        )
    ).all()
    for anterior in anteriores:
        anterior.ativo = False
        print(f"Plano alimentar {anterior.nome!r} desativado (continua no histórico).")

    plano = MealPlan(id=uuid7(), user_id=usuario.id, nome=nome_plano, ativo=True)
    session.add(plano)
    await session.flush()

    session.add(
        NutritionTarget(
            id=uuid7(),
            user_id=usuario.id,
            meal_plan_id=plano.id,
            proteina_g_kg=meta.proteina_g_kg,
            gordura_g_kg=meta.gordura_g_kg,
            fibra_g_por_1000kcal=meta.fibra_g_por_1000kcal,
            fator_atividade=meta.fator_atividade,
            ajuste_calorico=meta.ajuste_calorico,
            manutencao_kcal_manual=meta.manutencao_kcal_manual,
            sexo=meta.sexo,
            idade=meta.idade,
            altura_cm=meta.altura_cm,
        )
    )

    total_itens = 0
    for refeicao in refeicoes:
        slot = MealSlot(
            id=uuid7(),
            user_id=usuario.id,
            meal_plan_id=plano.id,
            nome=refeicao.nome,
            horario_alvo=_horario_sugerido(refeicao.nome),
            descricao=", ".join(refeicao.alimentos)[:240] or None,
            ordem=refeicao.ordem,
        )
        session.add(slot)
        await session.flush()

        for ordem, nome_alimento in enumerate(refeicao.alimentos):
            food_id = alimentos_por_nome.get(nome_alimento)
            if food_id is None:
                raise SystemExit(f"Alimento não encontrado no catálogo: {nome_alimento!r}")
            session.add(
                MealSlotItem(
                    id=uuid7(),
                    user_id=usuario.id,
                    meal_slot_id=slot.id,
                    food_item_id=food_id,
                    quantidade_g=None,
                    ordem=ordem,
                )
            )
            total_itens += 1

    print(
        f"Plano alimentar {nome_plano!r} criado para {usuario.email}: "
        f"{len(refeicoes)} refeições, {total_itens} itens."
    )


async def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--email", required=True, help="E-mail do usuário dono do plano.")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--plan-name", default=DEFAULT_PLAN_NAME)
    parser.add_argument(
        "--database-url",
        default=None,
        help="Postgres OWNER url. Por padrão, lê DATABASE_OWNER_URL do ambiente.",
    )
    parser.add_argument(
        "--force", action="store_true", help="Recria o plano se já existir um com este nome."
    )
    args = parser.parse_args()

    import os

    database_url = args.database_url or os.environ.get("DATABASE_OWNER_URL")
    if not database_url:
        raise SystemExit("Defina --database-url ou a variável DATABASE_OWNER_URL.")

    if not args.xlsx.exists():
        raise SystemExit(f"Planilha não encontrada: {args.xlsx}")

    dados = carregar_planilha(args.xlsx)

    # statement_cache_size=0: o pgbouncer do Neon (endpoint "-pooler") não
    # aceita prepared statement do asyncpg. O timeout largo é pelo cold start
    # do compute serverless, que passa do padrão do asyncpg e abortaria o seed
    # por impaciência — seed é raro e manual, esperar não custa nada.
    engine = create_async_engine(
        database_url,
        connect_args={"statement_cache_size": 0, "timeout": 180, "command_timeout": 180},
    )
    async with AsyncSession(engine) as session:
        usuario = await _obter_usuario(session, args.email)
        alimentos_por_nome = await _upsert_alimentos_globais(session, dados["alimentos"])
        await _upsert_suplementos_globais(session, dados["suplementos"])
        await session.flush()

        await _criar_plano(
            session,
            usuario,
            args.plan_name,
            dados["refeicoes"],
            alimentos_por_nome,
            dados["meta"],
            args.force,
        )
        await session.commit()

    await engine.dispose()
    print("Seed concluído.")


if __name__ == "__main__":
    asyncio.run(main())
